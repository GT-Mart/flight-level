# -*- coding: utf-8 -*-
import openpyxl
import os
import csv
import shutil
import datetime

from .log import get_logger

logger = None


def saveas(file_path, output_csv_path, archive_path, sheet_name=None, columns=None):
    # Open the Excel workbook
    wb = openpyxl.load_workbook(
        file_path, data_only=True
    )  # data_only=True to get values instead of formulas

    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    with open(output_csv_path, "w", newline="") as csv_file:
        writer = csv.writer(csv_file, quotechar='"', quoting=csv.QUOTE_ALL)

        # Write rows from the Excel sheet to the CSV file
        include_rows = False
        rows_added = False
        start_idx = 0
        for row in sheet.iter_rows():
            if row[0].value == "PLU No." or row[1].value == "PLU No.":
                include_rows = True
                if row[1].value == "PLU No." and row[0].value != "PLU No.":
                    start_idx = 1

            if include_rows:
                rows_added = True
                writer.writerow(
                    [
                        str(cell.value).replace("\n", " ") if cell.value else None
                        for cell in row[start_idx:]
                    ]
                )

        if not include_rows:
            writer.writerow(columns)
            for row in sheet.iter_rows():
                writer.writerow(
                    [
                        str(cell.value).replace("\n", " ") if cell.value else None
                        for cell in row[1:]
                    ]
                )
                rows_added = True

        if rows_added:
            shutil.move(file_path, archive_path)


def build_date(filename):
    global logger
    name_parts = filename.split("_")
    if len(name_parts) >= 3:
        sales_year = int(name_parts[0])
        sales_day = int(name_parts[1])
        sales_month = int(name_parts[2])

        if sales_month > 12:
            sales_day = int(name_parts[2])
            sales_month = int(name_parts[1])

        if sales_day > 31:
            sales_day = int(name_parts[1][0:2])
    else:
        logger.info("filename is not formatted correctly")
        return None

    return datetime.datetime(sales_year, sales_month, sales_day)


def run(config, job_name):
    global logger
    i = 0
    logger = get_logger(job_name, config)

    logger.info("Starting process to convert excel files to csv...")
    for file in os.listdir(config.PDF_SALES_FOLDER):
        filename, ext = os.path.splitext(file)
        if ext in config.EXCEL_EXTENSIONS:
            logger.info(file)
            sales_date = build_date(filename)
            saveas(
                os.path.join(config.PDF_SALES_FOLDER, file),
                os.path.join(
                    config.PDF_SALES_FOLDER,
                    f"{sales_date.year}_{sales_date.month}_{sales_date.day}_sales.csv",
                ),
                os.path.join(config.PDF_SALES_ARCHIVE_FOLDER, file),
                columns=config.EXCEL_COLUMNS,
            )
    logger.info("Process completed.")
